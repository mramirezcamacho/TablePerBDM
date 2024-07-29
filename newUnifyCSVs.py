import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill
import chardet

# Path to the folder containing CSV files
folder_path = 'BDL_tables'

# List all CSV files in the folder and sort them
csv_files = sorted([f for f in os.listdir(folder_path) if f.endswith('.csv')])

# Create a new Excel workbook
wb = Workbook()

# Define a border style
border_style = Border(left=Side(border_style="thin"),
                      right=Side(border_style="thin"),
                      top=Side(border_style="thin"),
                      bottom=Side(border_style="thin"))

# Define the bold font style
bold_font = Font(bold=True)
bold_big = Font(bold=True, size=14)

sheet_tab_colors = ['00ff00', 'ff9900', '0000ff', 'ffff00']

# Define the fill style for rows containing 'SME' or 'CKA'
highlight_hard_orange = PatternFill(
    start_color='FF9900', end_color='FF9900', fill_type='solid')
highlight_light_orange = PatternFill(
    start_color='F6b26b', end_color='F6b26b', fill_type='solid')
highlight_light_gray = PatternFill(
    start_color='F3f3f3', end_color='F3f3f3', fill_type='solid')
highlight_white = PatternFill(
    start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

for index, csv_file in enumerate(csv_files):
    # Extract sheet name from CSV file name (remove extension)
    sheet_name = os.path.splitext(csv_file)[0]

    # Create a new sheet in the workbook
    ws = wb.create_sheet(title=sheet_name[2:].replace('OKR', ' OKR'))

    if index < 3:
        ws.sheet_properties.tabColor = sheet_tab_colors[0]
    elif index < 8:
        ws.sheet_properties.tabColor = sheet_tab_colors[1]
    elif index < len(csv_files) - 1:
        ws.sheet_properties.tabColor = sheet_tab_colors[2]
    else:
        ws.sheet_properties.tabColor = sheet_tab_colors[3]

    # Open the CSV file and write its contents to the sheet
    with open(os.path.join(folder_path, csv_file), 'rb') as file:
        raw_data = file.read()
        result = chardet.detect(raw_data)
        encoding = result['encoding']

    # Use the detected encoding
    with open(os.path.join(folder_path, csv_file), 'r', newline='', encoding=encoding) as file:
        reader = csv.reader(file)
        for row in reader:
            ws.append(row)
    # Adjust column widths based on the maximum length of data in each column
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = (max_length + 2)  # Add a little extra space
        ws.column_dimensions[column_letter].width = adjusted_width

    # Apply borders and bold formatting to specific cells
    toggle_white = True
    for row_index, row in enumerate(ws.iter_rows(), start=1):
        highlight_row = None  # Flag to indicate if the row should be highlighted

        for cell in row:
            # Check if cell contains non-empty text
            if isinstance(cell.value, str) and cell.value.strip():
                cell.border = border_style
                if (cell.value in {'CKA', 'SME'}) or ('OKR' in cell.value) or ('Baseline' in cell.value) or ('Metric' in cell.value) or ('Target' in cell.value) or ('BDM' in cell.value) or ('SME' in cell.value) or ('CKA' in cell.value):
                    cell.font = bold_font
                if "OKR's" in cell.value:
                    cell.font = bold_big
                if cell.value == 'Metric' or cell.value == 'BDM':
                    highlight_row = 'Hard Orange'
                elif cell.value in {'CKA', 'SME'}:
                    highlight_row = 'Light Orange'

        # Apply highlight fill to the entire row if necessary
        if highlight_row == 'Hard Orange':
            fill = highlight_hard_orange
        elif highlight_row == 'Light Orange':
            fill = highlight_light_orange
        else:
            fill = highlight_white if toggle_white else highlight_light_gray
            toggle_white = not toggle_white

        for cell in row:
            try:
                if cell.value is not None and isinstance(cell.value, str):
                    cell.fill = fill
            except:
                pass

        adjusted_width = (max_length + 2)  # Add a little extra space
        ws.column_dimensions[column_letter].width = adjusted_width


if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Save the workbook
wb.save('NewAllBDLs.xlsx')

print(f'Combined Excel file "NewAllBDLs.xlsx" has been created successfully.')
